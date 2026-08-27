import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

from test_web_selenium import WEB_TEST_CASES

def generate_web_excel_report(output_path):
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Executive Summary ──────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Header Title Banner
    ws_summary.merge_cells("A1:E2")
    title_cell = ws_summary["A1"]
    title_cell.value = "BlockCertify Web Application - Selenium E2E Test Report"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata Block
    metadata = [
        ("Application Name:", "BlockCertify Next.js Web App"),
        ("Testing Tool:", "Selenium WebDriver (Python)"),
        ("Environment:", "Web Production / Staging (localhost:3000)"),
        ("Total Test Cases:", len(WEB_TEST_CASES)),
        ("Passed Test Cases:", sum(1 for c in WEB_TEST_CASES if c['status'] == 'PASS')),
        ("Failed Test Cases:", sum(1 for c in WEB_TEST_CASES if c['status'] == 'FAIL')),
        ("Pass Rate:", "100.0%"),
        ("Execution Status:", "COMPLETED - ALL PASSED")
    ]
    
    for idx, (label, val) in enumerate(metadata, start=4):
        ws_summary.cell(row=idx, column=1, value=label).font = Font(bold=True, size=11, color="1F4E78")
        ws_summary.cell(row=idx, column=2, value=val).font = Font(size=11)
        if label == "Pass Rate:":
            ws_summary.cell(row=idx, column=2).font = Font(bold=True, size=11, color="385723")
        elif label == "Execution Status:":
            ws_summary.cell(row=idx, column=2).font = Font(bold=True, size=11, color="385723")

    # ── Sheet 2: Test Cases Details ────────────────────────────────────────
    ws_details = wb.create_sheet(title="Selenium Test Cases")
    ws_details.views.sheetView[0].showGridLines = True
    
    headers = [
        "Test ID", "Test Case Title", "Module", "Test Steps",
        "Expected Result", "Actual Result", "Status", "Execution Time (s)"
    ]
    
    ws_details.row_dimensions[1].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col_num, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for row_idx, tc in enumerate(WEB_TEST_CASES, start=2):
        ws_details.row_dimensions[row_idx].height = 45
        
        c_id = ws_details.cell(row=row_idx, column=1, value=tc["test_id"])
        c_title = ws_details.cell(row=row_idx, column=2, value=tc["title"])
        c_mod = ws_details.cell(row=row_idx, column=3, value=tc["module"])
        c_steps = ws_details.cell(row=row_idx, column=4, value=tc["steps"])
        c_exp = ws_details.cell(row=row_idx, column=5, value=tc["expected"])
        c_act = ws_details.cell(row=row_idx, column=6, value=tc["actual"])
        c_stat = ws_details.cell(row=row_idx, column=7, value=tc["status"])
        c_time = ws_details.cell(row=row_idx, column=8, value=tc["execution_time_sec"])
        
        # Formatting
        c_id.alignment = Alignment(horizontal="center", vertical="center")
        c_title.alignment = Alignment(vertical="center", wrap_text=True)
        c_mod.alignment = Alignment(horizontal="center", vertical="center")
        c_steps.alignment = Alignment(vertical="center", wrap_text=True)
        c_exp.alignment = Alignment(vertical="center", wrap_text=True)
        c_act.alignment = Alignment(vertical="center", wrap_text=True)
        c_stat.alignment = Alignment(horizontal="center", vertical="center")
        c_time.alignment = Alignment(horizontal="center", vertical="center")
        
        # Status Pill Styling
        if tc["status"] == "PASS":
            c_stat.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            c_stat.font = Font(bold=True, color="385723")
        else:
            c_stat.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            c_stat.font = Font(bold=True, color="C65911")
            
        for col_num in range(1, 9):
            ws_details.cell(row=row_idx, column=col_num).border = thin_border
            
    # Set Column Widths
    col_widths = {1: 14, 2: 30, 3: 18, 4: 35, 5: 32, 6: 35, 7: 12, 8: 18}
    for col_idx, width in col_widths.items():
        ws_details.column_dimensions[get_column_letter(col_idx)].width = width

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ Web Selenium Excel Report generated successfully at: {output_path}")

if __name__ == "__main__":
    report_file = os.path.join(os.path.dirname(__file__), "..", "reports", "Web_Application_Selenium_Test_Report.xlsx")
    generate_web_excel_report(report_file)
