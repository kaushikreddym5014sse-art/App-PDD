import openpyxl
from openpyxl.utils import get_column_letter
import os

# Path to the Selenium test report Excel file
FILE_PATH = os.path.join(os.path.dirname(__file__), "Web_Application_Selenium_Test_Report.xlsx")

def add_description_column(ws, cat_idx):
    """Insert a Description column at the index where Category was removed.
    Populate each row with a concise description of the test case.
    """
    # Insert new column for Description at the former Category position
    ws.insert_cols(cat_idx)
    ws.cell(row=1, column=cat_idx, value="Description")
    # Determine column indices for optional fields used in description
    header = [cell.value for cell in ws[1]]
    # Try to locate a column that holds a readable test name or ID
    name_col = None
    if "Test Name" in header:
        name_col = header.index("Test Name") + 1
    elif "Test ID" in header:
        name_col = header.index("Test ID") + 1
    # Populate Description for each test case row
    for row in range(2, ws.max_row + 1):
        if name_col:
            test_name = ws.cell(row=row, column=name_col).value or ""
        else:
            test_name = ""
        # Build a generic but clear description
        desc = f"User attempts to execute '{test_name}'. Expected behavior is defined in the test steps. Actual result should match the expected outcome."
        ws.cell(row=row, column=cat_idx, value=desc)

def main():
    if not os.path.exists(FILE_PATH):
        raise FileNotFoundError(f"Excel report not found at {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active
    # Identify the 'Category' column (if present)
    header = [cell.value for cell in ws[1]]
    if "Category" in header:
        cat_idx = header.index("Category") + 1  # 1‑based index
        # Remove the Category column
        ws.delete_cols(cat_idx)
        # Add Description column where Category used to be
        add_description_column(ws, cat_idx)
    else:
        # If Category column missing, simply append Description at the end
        cat_idx = ws.max_column + 1
        add_description_column(ws, cat_idx)
    # Save back to the same file preserving existing formatting
    wb.save(FILE_PATH)
    print(f"Modified {FILE_PATH}: removed 'Category' and added 'Description' column.")

if __name__ == "__main__":
    main()
